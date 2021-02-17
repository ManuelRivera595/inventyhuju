from django.shortcuts import render, get_object_or_404, get_list_or_404
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.base import TemplateView
from django.urls import reverse, reverse_lazy
from django.shortcuts import redirect
from django.http import HttpResponse
from django.utils.html import strip_tags #excel
from openpyxl import Workbook  #excel
from reportlab.lib.units import cm #pdf
from reportlab.lib import colors #pdf
from reportlab.pdfgen import canvas #pdf
from io import BytesIO #pdf
from django.views.generic import View #pdf
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle #pdf
from .models import Page
from .forms import PageForm

#staff
class StaffRequiredMixin(object):
    """
    Usuario miembro del staff
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return redirect(reverse_lazy('admin:login'))
        return super(StaffRequiredMixin, self).dispatch(request, *args, **kwargs)
    

# Create your views here.
class PageListView(StaffRequiredMixin, ListView):
    model = Page

class PageDetailView(StaffRequiredMixin, DetailView):
    model = Page

class PageCreate(StaffRequiredMixin, CreateView):
    model = Page
    form_class = PageForm
    success_url = reverse_lazy('pages:pages')

class PageUpdate(StaffRequiredMixin, UpdateView):
    model = Page
    form_class = PageForm
    template_name_suffix = '_update_form'
    def get_success_url(self):
        return reverse_lazy('pages:update', args=[self.object.id]) + '?ok'

class PageDelete(DeleteView):
    model = Page
    success_url = reverse_lazy('pages:pages')

# buscador
class BuscarView(TemplateView):
    def post(self, request, *args, **kwargs):
        buscar=request.POST['buscalo']
        pages = Page.objects.filter(title__contains=buscar.upper())
        return render(request, 'pages/buscar.html',
                    {'pages':pages, 'page':True})

#Descarga excel
class ReporteExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las pages de nuestra base de datos
        pages = Page.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'INVENTARIO'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:H1')
        #Creamos los encabezados desde la celda B3 hasta la H1
        ws['B3'] = 'ID'
        ws['C3'] = 'PRODUCTO'
        ws['D3'] = 'DETALLE'
        ws['E3'] = 'CANT|UND'
        ws['F3'] = 'CANT|KG'
        ws['G3'] = 'PRECIO COMPRA'
        ws['H3'] = 'PRECIO VENTA'       
        cont=5
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for page in pages:
            ws.cell(row=cont,column=2).value = page.id
            ws.cell(row=cont,column=3).value = page.title
            ws.cell(row=cont,column=4).value = strip_tags(page.content)
            ws.cell(row=cont,column=5).value = page.order
            ws.cell(row=cont,column=6).value = page.decimal
            ws.cell(row=cont,column=7).value = page.pcompra
            ws.cell(row=cont,column=8).value = page.pventa
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#fin formato Excel

#Descarga PDF
class ReportePDF(View):  
     
    #def cabecera(self,pdf):
        #Utilizamos el archivo logo_django.png que está guardado en la carpeta media/imagenes
        #archivo_imagen = settings.MEDIA_ROOT+'/imagenes/logo_django.png'
        #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
        #pdf.drawImage(archivo_imagen, 40, 750, 120, 90,preserveAspectRatio=True)

    def tabla(self,pdf,y):
        #Creamos una tupla de encabezados para neustra tabla
        encabezados = ('ID', 'PRODUCTO', 'DETALLE', 'CANT', 'KG', 'P.C', 'P.V')
        #Creamos una lista de tuplas que van a contener a las personas
        detalles = [(page.id, page.title, strip_tags(page.content), page.order, page.decimal, page.pcompra, page.pventa) for page in Page.objects.all()]
        #Establecemos el tamaño de cada una de las columnas de la tabla
        detalle_orden = Table([encabezados] + detalles, colWidths=[1 * cm, 4 * cm, 6 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm,])
        #Aplicamos estilos a las celdas de la tabla
        detalle_orden.setStyle(TableStyle(
            [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]
        ))
        #Establecemos el tamaño de la hoja que ocupará la tabla 
        detalle_orden.wrapOn(pdf, 800, 600)
        #Definimos la coordenada donde se dibujará la tabla
        detalle_orden.drawOn(pdf, 60,y)              
         
    def get(self, request, *args, **kwargs):
        #Indicamos el tipo de contenido a devolver, en este caso un pdf
        response = HttpResponse(content_type='application/pdf')
        #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
        buffer = BytesIO()
        #Canvas nos permite hacer el reporte con coordenadas X y Y
        pdf = canvas.Canvas(buffer)
        #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
        #self.cabecera(pdf)
        #Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
        pdf.setFont("Helvetica", 16)
        #Dibujamos una cadena en la ubicación X,Y especificada
        pdf.drawString(230, 800, u"INVENTARIO YHUJU")
        pdf.setFont("Helvetica", 14)
        #pdf.drawString(200, 770, u"REPORTE DE PERSONAS")
        y = 700
        self.tabla(pdf, y)
        #Con show page hacemos un corte de página para pasar a la siguiente
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

#fin formato PDF
